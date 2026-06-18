import argparse
import json
import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Set
import warnings
from openpyxl import Workbook

from data.utils import get_sample_with_id, loadEvaluationSamples
from graph_builder import build_nlp_subgraph, build_baseline_subgraph
from helper import pretty_print
from pipeline_state import GraphState
import sys

use_baseline = "--baseline" in sys.argv or "-baseline" in sys.argv

if use_baseline:
    graph = build_baseline_subgraph(evaluate=True)
    graph_with_out_evaluation = build_baseline_subgraph(evaluate=False)
else:
    graph = build_nlp_subgraph(evaluate=True)
    graph_with_out_evaluation = build_nlp_subgraph(evaluate=False)

warnings.filterwarnings(
    "ignore",
    message=r"Pydantic serializer warnings:.*",
    category=UserWarning,
    module=r"pydantic\.main",
)

def _serialize_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False)


def _state_columns() -> List[str]:
    return list(GraphState.__annotations__.keys())

def _result_to_row(sample_id: int, result: Dict[str, Any], elapsed_s: float, columns: Iterable[str]) -> Dict[str, Any]:
    # Eclapsed time is the time taken for the graph to process the question and produce a result, which is a key metric for evaluation.
    row = {
        "sample_id": sample_id,
        "elapsed_seconds": round(elapsed_s, 3),
    }
    for key in columns:
        row[key] = _serialize_cell_value(result.get(key))
    return row

def _is_perfect_f1(score: float) -> bool:
    return abs(score - 1.0) < 1e-9


def _write_evaluation_workbook(
    rows: List[Dict[str, Any]],
    attention_rows: List[Dict[str, Any]],
    evaluated_samples: List[Dict[str, Any]],
    summary: Dict[str, Any],
    total_elapsed: float,
) -> Path:
    output_dir = Path("outputs")
    output_dir.mkdir(parents=True, exist_ok=True)

    file_name = f"evaluation_all_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = output_dir / file_name

    workbook = Workbook()
    detail_sheet = workbook.active
    if detail_sheet is None:
        raise RuntimeError("Failed to create worksheet.")
    detail_sheet.title = "evaluation_results"

    if rows:
        detail_headers = list(rows[0].keys())
        detail_sheet.append(detail_headers)
        for row in rows:
            detail_sheet.append([row.get(header, "") for header in detail_headers])

    summary_sheet = workbook.create_sheet("Summary")
    evaluated_ids_sorted = sorted(summary["evaluated_ids"])
    attention_ids_sorted = sorted(summary["attention_ids"])

    summary_sheet.append(["field", "value"])
    summary_sheet.append(["evaluated_sample_ids", ",".join(str(sample_id) for sample_id in evaluated_ids_sorted)])
    summary_sheet.append(["total_samples", summary["total"]])
    summary_sheet.append(["failed_samples", summary["failed"]])
    summary_sheet.append(["attention_sample_ids", ",".join(str(sample_id) for sample_id in attention_ids_sorted)])
    summary_sheet.append(["attention_count", len(attention_ids_sorted)])
    summary_sheet.append(["f1_score", round(summary["sum_score"] / summary["total"], 4) if summary["total"] else 0.0])
    summary_sheet.append(["average_precision", round(summary["sum_precision"] / summary["total"], 4) if summary["total"] else 0.0])
    summary_sheet.append(["average_recall", round(summary["sum_recall"] / summary["total"], 4) if summary["total"] else 0.0])
    summary_sheet.append(["total_elapsed_seconds", round(total_elapsed, 3)])

    summary_sheet.append([])
    summary_sheet.append(["sample_id", "question", "f1_score", "status"])
    for sample in sorted(evaluated_samples, key=lambda item: int(item["sample_id"])):
        summary_sheet.append([
            sample["sample_id"],
            sample["question"],
            sample["score"],
            sample["status"],
        ])

    attention_sheet = workbook.create_sheet("Attention Samples")
    if attention_rows:
        attention_headers = list(attention_rows[0].keys())
        attention_sheet.append(attention_headers)
        for row in attention_rows:
            attention_sheet.append([row.get(header, "") for header in attention_headers])
    else:
        attention_sheet.append(["message"])
        attention_sheet.append(["No failed sample and all f1_score are 1.0"])

    workbook.save(file_path)
    return file_path

def _print_realtime_summary(completed: int, total: int, summary: Dict[str, Any]) -> None:
    avg_score = summary["sum_score"] / completed if completed else 0.0
    avg_precision = summary["sum_precision"] / completed if completed else 0.0
    avg_recall = summary["sum_recall"] / completed if completed else 0.0
    print(
        (
            f"[Progress] {completed}/{total} | "
            f"f1_score={avg_score:.4f} | "
            f"avg_precision={avg_precision:.4f} | "
            f"avg_recall={avg_recall:.4f} | "
            f"failed={summary['failed']}"
        ),
        flush=True,
    )


def _parse_eval_ids(eval_arg: str, max_id: int) -> List[int]:
    raw = (eval_arg or "").strip().lower()
    if raw == "all":
        return list(range(1, max_id + 1))

    selected: Set[int] = set()
    tokens = [token.strip() for token in raw.split(",") if token.strip()]
    if not tokens:
        raise ValueError("-eval must be one of: all, 3, 1-5, or 1,3,5")

    for token in tokens:
        if "-" in token:
            parts = [part.strip() for part in token.split("-")]
            if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit():
                raise ValueError(f"Invalid range '{token}'. Use format start-end, e.g. 1-5")
            start_id = int(parts[0])
            end_id = int(parts[1])
            if start_id > end_id:
                raise ValueError(f"Invalid range '{token}': start must be <= end")
            selected.update(range(start_id, end_id + 1))
            continue

        if not token.isdigit():
            raise ValueError(f"Invalid id '{token}'. Use integer ids like 1,3,5")
        selected.add(int(token))

    invalid_ids = sorted(sample_id for sample_id in selected if sample_id < 1 or sample_id > max_id)
    if invalid_ids:
        raise ValueError(f"Requested ids out of range 1..{max_id}: {invalid_ids}")

    return sorted(selected)


def evaluate_all_samples(selected_ids: Optional[List[int]] = None, workers: Optional[int] = None) -> None:
    samples = loadEvaluationSamples()
    max_id = len(samples)

    if max_id == 0:
        print("No evaluation samples found.")
        return

    if selected_ids is None:
        selected_ids = list(range(1, max_id + 1))

    selected_samples = [(sample_id, samples[sample_id - 1]) for sample_id in selected_ids]
    sample_lookup = {sample_id: sample for sample_id, sample in selected_samples}
    total = len(selected_samples)
    if total == 0:
        print("No sample ids selected for evaluation.")
        return

    default_workers = max(1, os.cpu_count() or 1)
    max_workers = default_workers if workers is None else workers
    print(f"Start parallel evaluation: total={total}, workers={max_workers}")

    columns = _state_columns()
    rows: List[Dict[str, Any]] = []
    attention_rows: List[Dict[str, Any]] = []
    evaluated_samples: List[Dict[str, Any]] = []
    summary = {
        "sum_score": 0.0,
        "sum_precision": 0.0,
        "sum_recall": 0.0,
        "failed": 0,
        "evaluated_ids": set(),
        "attention_ids": set(),
        "total": total,
    }

    start_all = time.perf_counter()

    def run_single(sample_id: int, sample: Mapping[str, Any]) -> Dict[str, Any]:
        start = time.perf_counter()
        result = graph.invoke({"question": sample["question"], "expected_nodes": sample["nodes"]})
        elapsed = time.perf_counter() - start
        return {
            "sample_id": sample_id,
            "result": result,
            "elapsed": elapsed,
        }

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(run_single, sample_id, sample): sample_id
            for sample_id, sample in selected_samples
        }

        completed = 0
        for future in as_completed(future_map):
            sample_id = future_map[future]
            completed += 1
            sample = sample_lookup[sample_id]
            question = str(sample.get("question", ""))

            try:
                item = future.result()
                result = item["result"]
                eval_info = result.get("evaluation", {}) if isinstance(result, dict) else {}
                score = float(eval_info.get("score", 0.0) or 0.0)
                eval_precision = float(eval_info.get("precision", 0.0) or 0.0)
                eval_recall = float(eval_info.get("recall", 0.0) or 0.0)
                has_error = bool(result.get("errors"))

                summary["sum_score"] += score
                summary["sum_precision"] += eval_precision
                summary["sum_recall"] += eval_recall
                summary["evaluated_ids"].add(sample_id)
                if has_error:
                    summary["failed"] += 1

                needs_attention = has_error or not _is_perfect_f1(score)
                if needs_attention:
                    summary["attention_ids"].add(sample_id)

                row = _result_to_row(item["sample_id"], result, item["elapsed"], columns)
                row["question"] = question
                row["f1_score"] = score
                row["precision"] = eval_precision
                row["recall"] = eval_recall
                row["has_errors"] = has_error
            except Exception as exc:
                summary["failed"] += 1
                summary["evaluated_ids"].add(sample_id)
                summary["attention_ids"].add(sample_id)
                score = 0.0
                row = _result_to_row(
                    sample_id=sample_id,
                    result={"errors": [str(exc)]},
                    elapsed_s=0.0,
                    columns=columns,
                )
                row["question"] = question
                row["f1_score"] = score
                row["precision"] = 0.0
                row["recall"] = 0.0
                row["has_errors"] = True

            rows.append(row)
            if bool(row.get("has_errors")) or not _is_perfect_f1(float(row.get("f1_score", 0.0) or 0.0)):
                attention_rows.append(row)

            evaluated_samples.append(
                {
                    "sample_id": sample_id,
                    "question": question,
                    "score": float(row.get("f1_score", 0.0) or 0.0),
                    "status": "needs_attention"
                    if bool(row.get("has_errors")) or not _is_perfect_f1(float(row.get("f1_score", 0.0) or 0.0))
                    else "ok",
                }
            )

            _print_realtime_summary(completed, total, summary)

    total_elapsed = time.perf_counter() - start_all
    output_file = _write_evaluation_workbook(
        rows=rows,
        attention_rows=attention_rows,
        evaluated_samples=evaluated_samples,
        summary=summary,
        total_elapsed=total_elapsed,
    )

    avg_score = summary["sum_score"] / total if total else 0.0
    avg_precision = summary["sum_precision"] / total if total else 0.0
    avg_recall = summary["sum_recall"] / total if total else 0.0

    print("\n=== Final Summary ===")
    print(f"Total samples: {total}")
    print(f"Failed samples: {summary['failed']}")
    print(f"F1 score: {avg_score:.4f}")
    print(f"Average precision: {avg_precision:.4f}")
    print(f"Average recall: {avg_recall:.4f}")
    print(f"Total elapsed seconds: {total_elapsed:.3f}")
    print(f"Excel output: {output_file}")

def main() -> None:
    parser = argparse.ArgumentParser(description="DoctorChatbot runner")
    parser.add_argument(
        "-eval",
        type=str,
        help="Evaluation ids: all | 3 | 1-5 | 1,3,5",
        default=None,
    )
    parser.add_argument("-id", type=int, default=11, help="Sample id for single run mode")
    parser.add_argument(
        "-q",
        "--question",
        type=str,
        default=None,
        help="Ask any question directly",
    )
    parser.add_argument(
        "-workers",
        "--workers",
        type=int,
        default=None,
        help="Number of worker threads for evaluation mode. Default is cpu_count.",
    )

    parser.add_argument(
        "-debug",
        "--debug",
        action="store_true",
        help="Enable debug mode with more verbose output",
    )

    parser.add_argument(
        "--baseline",
        action="store_true",
        help="Use the LLM-only schema baseline",
    )

    args = parser.parse_args()

    if args.eval is not None:
        if args.workers is not None and args.workers < 1:
            parser.error("-workers must be >= 1")
            return

        total_samples = len(loadEvaluationSamples())
        try:
            selected_ids = _parse_eval_ids(args.eval, total_samples)
        except ValueError as exc:
            parser.error(str(exc))
            return

        evaluate_all_samples(selected_ids, workers=args.workers)
        return

    if args.question is not None:
        result = graph_with_out_evaluation.invoke({"question": args.question})
        pretty_print(result)
        return
    
    if args.debug:
        from config import settings
        settings.DEBUG = True

    sample = get_sample_with_id(args.id)
    result = graph.invoke({"question": sample["question"], "expected_nodes": sample["nodes"]})
    pretty_print(result)

if __name__ == "__main__":
    main()
    
